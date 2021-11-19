import os
from datetime import date, datetime

from django.shortcuts import redirect
from django.http import HttpResponseRedirect, HttpResponse, Http404, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import PatternFill

from .models import Campus, Segmento, Curso, Pergunta, ParticipacaoPergunta, RespostaObjetiva, \
    Atuacao, Lotacao, Questionario
from django.db import connection
import json
from django.http import JsonResponse
from django.conf import settings


def index(request):
    return redirect('http://cpa.bcc.anapolis.ifg.edu.br')


@csrf_exempt
def apianswer(request):
    segmento = ''
    skey = request.GET.get('skey')
    if skey == "s4UkHMQC":
        segmento = "Estudante"
    elif skey == "zc3WsGum":
        segmento = "Docente"
    elif skey == "g3YTAfpT":
        segmento = "Servidor Técnico"
    elif skey == "4jn7qduk":
        segmento = "Servidor Técnico da Reitoria"

    if request.method == "GET":
        # campuses = [{'id': row['id'], 'nome': row['nome']} for row in Campus.objects.all().order_by('nome').values('id', 'nome')]

        ultimo_ano = Questionario.objects.order_by('-ano').filter(perguntas_publico__exact=True).values('id',
                                                                                                        'ano').first()

        perguntas = [{'id': pergunta['id'],
                      'titulo': pergunta['titulo'],
                      'tipo': pergunta['tipo'],
                      'lotacao': pergunta['perguntasegmento__lotacao__titulo'],
                      'dimensao': pergunta['dimensao__dimensao'],
                      'eixo': pergunta['dimensao__eixo__eixo']} for pergunta in
                     Pergunta.objects.filter(perguntasegmento__segmento__nome=segmento).filter(
                         dimensao__eixo__questionario=ultimo_ano['id']).order_by("dimensao__eixo").order_by(
                         "dimensao__eixo").order_by(
                         "tipo").values('id', 'titulo', 'tipo', 'dimensao__dimensao', 'dimensao__eixo__eixo',
                                        'perguntasegmento__lotacao__titulo',
                                        'dimensao__eixo__questionario').distinct()]

        resp_objetivas = [{'id': pergunta['id'], 'titulo': pergunta['titulo'], 'value': pergunta['value']} for
                          pergunta
                          in RespostaObjetiva.objects.all().order_by("-value").values('id', 'titulo', 'value')]

        segmento = Segmento.objects.get(nome=segmento)

        eixos = {}

        for pergunta in perguntas:
            eixo = pergunta['eixo']
            dimensao = pergunta['dimensao']

            if eixo not in eixos:
                eixos[eixo] = {}
            if dimensao not in eixos[eixo]:
                eixos[eixo][dimensao] = []

            pergunta.pop('eixo', None)
            pergunta.pop('dimensao', None)

            eixos[eixo][dimensao].append(pergunta)

        return JsonResponse({
            "segmento": {'id': segmento.id, 'nome': segmento.nome},
            # "campus": campuses,
            "ano": ultimo_ano['ano'],
            "perguntas": eixos,
            "resp_objetivas": resp_objetivas
        })

    elif request.method == 'POST':
        now = datetime.now()
        start_date = datetime.strptime(os.getenv('QUEST_START_DATE'), '%Y-%m-%d')
        end_date = datetime.strptime(os.getenv('QUEST_END_DATE'), '%Y-%m-%d')

        if now < start_date or now > end_date:
            return HttpResponse(status=405)

        body_unicode = request.body.decode('utf-8')
        body = json.loads(body_unicode)
        ParticipacaoPergunta.create_participacao(atuacao=body['atuacao'], lotacao=body['lotacao'],
                                                 segmento=body['segmento'], curso=body['curso'], campus=body['campus'],
                                                 perguntas=body['respostas'], ano=body['ano'])
        return JsonResponse({
            "content": 0,
        })


def apiano(request):
    with connection.cursor() as cursor:
        cursor.execute('select distinct ano from informacoes order by ano desc')
        anos = [int(row[0]) for row in cursor.fetchall()]
        return JsonResponse({
            'ano': anos
        })


def apiatuacao(request):
    pergunta = request.GET.get('pergunta', None)
    campus = request.GET.get("campus", None)
    segmento = request.GET.get("segmento", None)
    ano = request.GET.get("ano", None)
    eixo = request.GET.get("eixo", None)
    dimensao = request.GET.get("dimensao", None)
    atuacao = []

    if pergunta is None:
        atuacao = [{'id': atuacao['id'], 'nome': atuacao['titulo']} for atuacao in
                   Atuacao.objects.values('id', 'titulo').all()]
        return JsonResponse({"atuacao": atuacao})

    sql = 'select distinct atuacao_id, atuacao from informacoes where pergunta_id = {} and ano = {} '.format(
        int(pergunta), int(ano))

    if campus is not None:
        sql += 'and campus_id = {} '.format(int(campus))
    if segmento is not None:
        sql += 'and segmento_id = {} '.format(int(segmento))
    if dimensao is not None and dimensao != "0":
        sql += ' and dimensao_id = {} '.format(int(dimensao))
    if eixo is not None and eixo != "0":
        sql += ' and eixo_id = {} '.format(int(eixo))

    sql += 'and atuacao IS NOT NULL order by atuacao'

    with connection.cursor() as cursor:
        cursor.execute(sql)
        atuacao = [{'id': row[0], 'nome': row[1]} for row in cursor.fetchall()]

    return JsonResponse({"atuacao": atuacao})


def apilotacao(request):
    pergunta = request.GET.get('pergunta', None)
    campus = request.GET.get("campus", None)
    segmento = request.GET.get("segmento", None)
    ano = request.GET.get("ano", None)
    eixo = request.GET.get("eixo", None)
    dimensao = request.GET.get("dimensao", None)
    lotacao = []

    if pergunta is None:
        lotacao = [{'id': lotacao['id'], 'nome': lotacao['titulo']} for lotacao in
                   Lotacao.objects.values('id', 'titulo').all()]
        return JsonResponse({"lotacao": lotacao})

    sql = 'select distinct lotacao_id, lotacao from informacoes where pergunta_id = {} and ano = {} '.format(
        int(pergunta), int(ano))

    if campus is not None:
        sql += 'and campus_id = {} '.format(int(campus))
    if segmento is not None:
        sql += 'and segmento_id = {} '.format(int(segmento))
    if dimensao is not None and dimensao != "0":
        sql += ' and dimensao_id = {} '.format(int(dimensao))
    if eixo is not None and eixo != "0":
        sql += ' and eixo_id = {} '.format(int(eixo))

    sql += 'and lotacao IS NOT NULL order by lotacao'

    with connection.cursor() as cursor:
        cursor.execute(sql)
        lotacao = [{'id': row[0], 'nome': row[1]} for row in cursor.fetchall()]

    return JsonResponse({"lotacao": lotacao})


def apicurso(request):
    pergunta = request.GET.get("pergunta", None)
    campus = request.GET.get("campus", None)
    segmento = request.GET.get("segmento", None)
    ano = request.GET.get("ano", None)
    eixo = request.GET.get("eixo", None)
    dimensao = request.GET.get("dimensao", None)

    cursos = []

    if pergunta is None:
        cursos = [{'id': curso['id'], 'nome': curso['nome']} for curso in
                  Curso.objects.filter(cursocampus__campus_id=campus).exclude(nome='Não Aplica').order_by(
                      'nome').values(
                      'nome', 'id')]
    else:
        sql = 'select distinct curso_id, curso from informacoes where curso != \'Não Aplica\' and pergunta_id = {} and ano = {} '.format(
            int(pergunta), int(ano))
        if campus is not None:
            sql += 'and campus_id = {} '.format(int(campus))
        if segmento is not None:
            sql += 'and segmento_id = {} '.format(int(segmento))
        if dimensao is not None and dimensao != "0":
            sql += ' and dimensao_id = {} '.format(int(dimensao))
        if eixo is not None and eixo != "0":
            sql += ' and eixo_id = {} '.format(int(eixo))
        sql += 'order by curso'

        with connection.cursor() as cursor:
            cursor.execute(sql)
            cursos = [{'id': row[0], 'nome': row[1]} for row in cursor.fetchall()]
    return JsonResponse({"cursos": cursos})


def apicampus(request):
    pergunta = request.GET.get("pergunta", None)
    segmento = request.GET.get("segmento", None)
    ano = request.GET.get("ano", None)
    eixo = request.GET.get("eixo", None)
    dimensao = request.GET.get("dimensao", None)
    campus = []

    if pergunta is None:  # LISTAR TODOS OS CAMPUS
        campus = [{'id': row['id'], 'nome': row['nome']} for row in
                  Campus.objects.all().order_by('nome').values('id', 'nome')]

        return JsonResponse({'campus': campus})

    sql = 'select distinct campus, campus_id from informacoes where ano = {} '.format(int(ano))

    if pergunta != '0':
        sql += ' and pergunta_id = {}'.format(int(pergunta))
    if segmento is not None:
        sql += ' and segmento_id = {} '.format(int(segmento))
    if dimensao is not None and dimensao != "0":
        sql += ' and dimensao_id = {} '.format(int(dimensao))
    if eixo is not None and eixo != "0":
        sql += ' and eixo_id = {} '.format(int(eixo))

    sql += ' order by campus;'

    with connection.cursor() as cursor:
        cursor.execute(sql)
        campus = [{'id': row[1], 'campus': row[0]} for row in cursor.fetchall()]

    return JsonResponse({'campus': campus})


def apiupdate(request):
    with connection.cursor() as cursor:
        cursor.execute('refresh materialized view informacoes')
        cursor.close()
    return HttpResponseRedirect('http://cpa.bcc.anapolis.ifg.edu.br/relatorio')


def apipergunta(request):
    ano = request.GET.get("ano", None)
    eixo = request.GET.get("eixo", None)
    dimensao = request.GET.get("dimensao", None)

    if ano is not None:
        sql = 'select distinct pergunta_id, pergunta from informacoes where pergunta not like \'Caso%\' and ano = {}'.format(
            int(ano))

        if dimensao is not None and dimensao != "0":
            sql += ' and dimensao_id = {} '.format(int(dimensao))
        if eixo is not None and eixo != "0":
            sql += ' and eixo_id = {} '.format(int(eixo))

        sql += ' order by pergunta'

        with connection.cursor() as cursor:
            cursor.execute(sql)
            data = [{'id': row[0], 'titulo': row[1]} for row in cursor.fetchall()]

    return JsonResponse({'dados': data})


def apigrafico(request):
    pergunta = request.GET.get("pergunta", None)
    segmento = request.GET.get("segmento", None)
    atuacao = request.GET.get("atuacao", None)
    lotacao = request.GET.get("lotacao", None)
    dimensao = request.GET.get("dimensao", None)
    eixo = request.GET.get("eixo", None)
    campus = request.GET.get("campus", None)
    curso = request.GET.get("curso", None)
    ano = request.GET.get("ano", None)
    data = []
    segmentos = []
    respostas = []

    if pergunta != "0":
        sql = 'select count(pessoa), segmento, resposta, resposta_id from informacoes where '
        sql += 'pergunta_id = {} '.format(int(pergunta))
        if segmento is not None:
            sql += ' and segmento_id = {} '.format(int(segmento))
        if ano is not None:
            sql += ' and ano = {} '.format(int(ano))
        if dimensao is not None and dimensao != "0":
            sql += ' and dimensao_id = {} '.format(int(dimensao))
        if eixo is not None and eixo != "0":
            sql += ' and eixo_id = {} '.format(int(eixo))
        if campus is not None:
            sql += ' and campus_id = {} '.format(int(campus))
        if curso is not None:
            sql += ' and curso_id = {} '.format(int(curso))
        if atuacao is not None:
            sql += ' and atuacao_id = {} '.format(int(atuacao))
        if lotacao is not None:
            sql += ' and lotacao_id = {} '.format(int(lotacao))

        sql += ' group by segmento, resposta, resposta_id order by resposta_id;'
        total = 0
        indicador = 0
        # print(sql)
        with connection.cursor() as cursor:
            cursor.execute(sql)
            graficos = cursor.fetchall()
            data = [{'count': row[0], 'segmento': row[1], 'resposta': row[2]} for row in graficos]
            for row in graficos:
                total += row[0]
                if row[1] not in segmentos:
                    segmentos.append(row[1])
                if row[2] not in respostas:
                    respostas.append(row[2])
                if row[2] == 'Ótimo' or row[2] == 'Bom':
                    indicador += row[0]
            total = round((indicador * 100 / total) * 10) / 10
        if total > 75:
            indicador = {'label': 'Manter', 'valor': total, 'cor': '#008ffb'}
        elif total > 50:
            indicador = {'label': 'Desenvolver', 'valor': total, 'cor': '#00e396'}
        elif total > 25:
            indicador = {'label': 'Melhorar', 'valor': total, 'cor': '#feb019'}
        else:
            indicador = {'label': 'Sanar', 'valor': total, 'cor': '#ff4560'}
        pergunta = Pergunta.objects.get(pk=pergunta).titulo
        return JsonResponse(
            {'pergunta': pergunta, 'indicador': indicador, 'roles': segmentos, 'respostas': respostas,
             'data': data})
    else:  # grafico em pizza
        sql = 'select count(distinct pessoa), campus'
        fetch = ''
        group = ' group by campus'
        where = ' where ano = {}'.format(int(ano))

        if dimensao is not None and dimensao != "0":
            where += ' and dimensao_id = {} '.format(int(dimensao))
        if eixo is not None and eixo != "0":
            where += ' and eixo_id = {} '.format(int(eixo))

        if campus is not None and segmento is not None:
            nome_segmento = Segmento.objects.get(pk=segmento).nome
            fetch = ', segmento'
            group += ', segmento'
            where += ' and segmento_id = {} and campus_id = {} '.format(int(segmento), int(campus))
            # print(nome_segmento)
            if nome_segmento == "Docente":
                fetch += ', atuacao '
                group += ', atuacao '
            elif nome_segmento == "Estudante":
                fetch += ', curso '
                group += ', curso '
            else:
                fetch += ', lotacao '
                group += ', lotacao '
        elif segmento is not None:
            fetch = ', segmento'
            group += ', segmento'
            where += ' and segmento_id = {}'.format(int(segmento))
        elif campus is not None:
            fetch = ', segmento'
            group += ', segmento'
            where += ' and campus_id = {}'.format(int(campus))

        sql = sql + fetch + ' from informacoes ' + where + group + ';'

        with connection.cursor() as cursor:
            cursor.execute(sql)
            graficos = cursor.fetchall()
            # print(sql, graficos)
        for row in graficos:
            if campus is not None and segmento is not None:
                data.append({'count': row[0], 'label': row[3]})
                segmentos.append(row[3])
            elif campus is not None:
                data.append({'count': row[0], 'label': row[2]})
                segmentos.append(row[2])
            else:
                data.append({'count': row[0], 'label': row[1]})
                segmentos.append(row[1])

        return JsonResponse({'roles': segmentos, 'data': data})


def apisegmento(request):
    pergunta = request.GET.get('pergunta', None)
    ano = request.GET.get('ano', None)
    eixo = request.GET.get("eixo", None)
    dimensao = request.GET.get("dimensao", None)
    segmentos = []
    if pergunta is not None:
        sql = 'select distinct segmento_id, segmento from informacoes where ano = {} '.format(int(ano))

        if pergunta != '0':
            sql += ' and pergunta_id = {} '.format(int(pergunta))
        if dimensao is not None and dimensao != "0":
            sql += ' and dimensao_id = {} '.format(int(dimensao))
        if eixo is not None and eixo != "0":
            sql += ' and eixo_id = {} '.format(int(eixo))

        sql += ' order by segmento;'

        # print(sql)

        with connection.cursor() as cursor:
            cursor.execute(sql)
            segmentos = [{'id': row[0], 'nome': row[1]} for row in cursor.fetchall()]

    return JsonResponse({'segmentos': segmentos})


def apieixo(request):
    ano = request.GET.get('ano', None)
    eixos = []
    if ano is not None:
        with connection.cursor() as cursor:
            cursor.execute(
                'select distinct eixo_id, eixo from informacoes where ano = {} order by eixo'.format(int(ano)))
            eixos = [{'id': row[0], 'nome': row[1]} for row in cursor.fetchall() if row[0] is not None]
    return JsonResponse({'eixos': eixos})


def apidimensao(request):
    ano = request.GET.get('ano', None)
    eixo = request.GET.get('eixo', None)

    dimensoes = []

    if ano is not None and eixo is not None:
        with connection.cursor() as cursor:
            cursor.execute(
                'select distinct dimensao_id, dimensao from informacoes where ano = {} and eixo_id = {} order by dimensao'.format(
                    int(ano), int(eixo)))
            dimensoes = [{'id': row[0], 'nome': row[1]} for row in cursor.fetchall() if row[0] is not None]

    return JsonResponse({'dimensoes': dimensoes})


def download_excel_file(path):
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404


def apirelatorio(request):
    try:
        questionario = int(request.GET.get('questionario'))
    except TypeError:
        return HttpResponseBadRequest

    from openpyxl import Workbook
    from openpyxl.styles import Font, Color, Fill
    from openpyxl.utils import get_column_letter
    wb = Workbook()

    dest_filename = f'graph/static/relatorio_{questionario}.xlsx'

    if os.path.exists(dest_filename):
        return download_excel_file(dest_filename)

    with connection.cursor() as cursor:
        cursor.execute(
            """select distinct ano from public.informacoes where ano = %d;""" % questionario)
        anos = [ano[0] for ano in cursor.fetchall()]

        if questionario not in anos:
            raise Http404

    with connection.cursor() as cursor:
        cursor.execute(
            """select distinct pergunta, pergunta_id from public.informacoes where ano = %d order by pergunta_id;""" % questionario)
        perguntas = {}
        for pergunta in cursor.fetchall():
            perguntas[pergunta[0]] = pergunta[1]

    with connection.cursor() as cursor:
        cursor.execute(
            """select distinct campus, campus_id from public.informacoes where ano = %d;""" % questionario)
        campuses = {}
        for campus in cursor.fetchall():
            campuses[campus[0]] = campus[1]

    spreadsheet = {}

    with connection.cursor() as cursor:
        cursor.execute(
            """select pessoa, segmento, curso, lotacao, atuacao, pergunta_id, resposta, campus_id, subjetiva from public.informacoes where ano = %d order by pergunta_id""" % questionario)

        for row in cursor.fetchall():
            pessoa = row[0]
            segmento = row[1]
            curso = row[2]
            lotacao = row[3]
            atuacao = row[4]
            pergunta_id = row[5]
            resposta = row[6]
            campus_id = row[7]

            if row[6] is None and row[8] is not None:
                resposta = row[8]

            if campus_id not in spreadsheet:
                spreadsheet[campus_id] = {}

            if pessoa not in spreadsheet[campus_id]:
                spreadsheet[campus_id][pessoa] = {
                    'pessoa': pessoa,
                    'segmento': segmento,
                    'curso': curso if curso != 'Não Aplica' else '',
                    'lotacao': lotacao,
                    'atuacao': atuacao,
                    'perguntas': {}
                }

            spreadsheet[campus_id][pessoa]['perguntas'][pergunta_id] = resposta

    spreadsheet_perguntas = {}

    for i, pergunta in enumerate(list(perguntas.values())):
        spreadsheet_perguntas[pergunta] = i

    for index, campus in enumerate(list(campuses.keys())):
        if index == 0:
            wb_current = wb.active
            wb_current.title = campus
        else:
            wb_current = wb.create_sheet(campus)

        ft = Font(color='ffffff', bold=True)
        fill = PatternFill('solid', fgColor='000000')

        for row in wb_current.iter_rows(min_row=1, max_row=1, min_col=1, max_col=len(list(perguntas.keys())) + 5):
            for cell in row:
                cell.fill = fill
                cell.font = ft

        wb_current['A1'] = 'ID pessoa'
        wb_current['B1'] = 'Segmento'
        wb_current['C1'] = 'Curso'
        wb_current['D1'] = 'Lotação'
        wb_current['E1'] = 'Area de Atuação'

        for i, pergunta in enumerate(list(perguntas.values())):
            wb_current[f'{get_column_letter(spreadsheet_perguntas[pergunta] + 6)}1'] = list(perguntas.keys())[i]

        for i, pessoa in enumerate(list(spreadsheet[campuses[campus]].keys())):
            wb_current[f'A{i + 2}'] = spreadsheet[campuses[campus]][pessoa]['pessoa']
            wb_current[f'B{i + 2}'] = spreadsheet[campuses[campus]][pessoa]['segmento']
            wb_current[f'C{i + 2}'] = spreadsheet[campuses[campus]][pessoa]['curso']
            wb_current[f'D{i + 2}'] = spreadsheet[campuses[campus]][pessoa]['lotacao']
            wb_current[f'E{i + 2}'] = spreadsheet[campuses[campus]][pessoa]['atuacao']

            for j, pergunta in enumerate(list(spreadsheet[campuses[campus]][pessoa]['perguntas'].keys())):
                wb_current[f'{get_column_letter(spreadsheet_perguntas[pergunta] + 6)}{i + 2}'] = \
                    spreadsheet[campuses[campus]][pessoa]['perguntas'][pergunta]

    wb.save(dest_filename)

    return download_excel_file(dest_filename)
